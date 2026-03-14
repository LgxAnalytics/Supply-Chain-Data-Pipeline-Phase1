import os
import shutil
import pdfplumber
import pandas as pd
import re
import logging
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# 1. DYNAMIC ENVIRONMENT SETUP
base_path = base_path = os.path.dirname(os.path.abspath(__file__))
archive_base = os.path.join(base_path, "ARCHIVE")
reports_base = os.path.join(base_path, "Incoming_Reports")
history_file = os.path.join(base_path, "processed_history.txt")
log_file = os.path.join(base_path, "system_audit.log")

# Auto-provisioning of directory structure
for folder in [archive_base, reports_base]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# 2. LOGGING INITIALIZATION
now = datetime.now()
current_week = now.isocalendar()[1]
output_file = os.path.join(reports_base, f"Incoming_Inbound_Week_{current_week}.xlsx")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.FileHandler(log_file), logging.StreamHandler()]
)

logging.info(f"=== SYSTEM START: PIPELINE WEEK {current_week} ===")
logging.info(f"Root Directory: {base_path}")

# Load Duplicate Shield history
if os.path.exists(history_file):
    with open(history_file, "r") as f:
        processed_files = set(f.read().splitlines())
else:
    processed_files = set()

# Scan for new PDF files in current folder
all_pdf_files = [f for f in os.listdir(base_path) if f.endswith('.pdf')]
new_pdf_files = [f for f in all_pdf_files if f not in processed_files]

if not new_pdf_files:
    logging.info("Status: No new PDF files found in the root folder.")
else:
    logging.info(f"Workload: Detected {len(new_pdf_files)} new file(s).")
    extracted_data = []

    # 3. EXTRACTION ENGINE
    for file_name in new_pdf_files:
        logging.info(f"Analyzing: {file_name}")
        current_po = "NO_PO_FOUND"
        file_full_path = os.path.join(base_path, file_name)
        
        try:
            creation_timestamp = os.path.getctime(file_full_path)
            download_date = datetime.fromtimestamp(creation_timestamp).strftime('%Y-%m-%d %H:%M')

            with pdfplumber.open(file_full_path) as pdf:
                full_text = ""
                for page in pdf.pages:
                    text = page.extract_text()
                    if text: full_text += text + "\n"
                
                po_match = re.search(r'PO\d+', full_text)
                if po_match: current_po = po_match.group()

                lines = full_text.split('\n')
                items_in_file = 0

                for line in lines:
                    parts = line.strip().split()
                    if len(parts) < 4 or not re.match(r'^\d+$', parts[0]):
                        continue

                    sku_idx = -1
                    sku = None
                    
                    for i, part in enumerate(parts[1:], start=1):
                        if re.match(r'^[A-Z0-9_-]{6,25}$', part) and any(c.isdigit() for c in part):
                            sku_idx, sku = i, part
                            break
                    
                    if sku_idx != -1 and not sku.startswith('QA'):
                        desc = " ".join(parts[sku_idx + 1:])
                        qty = next((p for p in parts[1:sku_idx] if p.isdigit()), None)
                        
                        if qty and desc:
                            sku_clean = re.sub(r'_[A-Z0-9]+$', '', sku)
                            extracted_data.append({
                                "Download_Date": download_date,
                                "PO_Number": current_po,
                                "Item_SKU": sku_clean,
                                "Quantity": int(qty),
                                "Description": desc,
                                "Source_File": file_name
                            })
                            items_in_file += 1
            
            logging.info(f"Success: Extracted {items_in_file} lines from {file_name}.")
            with open(history_file, "a") as f: f.write(file_name + "\n")

        except Exception as e:
            logging.error(f"Critical error processing {file_name}: {e}")

    # 4. EXCEL EXPORT & PREMIUM FORMATTING
    if extracted_data:
        if os.path.exists(output_file):
            existing_df = pd.read_excel(output_file)
            df = pd.concat([existing_df, pd.DataFrame(extracted_data)], ignore_index=True)
            logging.info("Appending data to existing weekly report.")
        else:
            df = pd.DataFrame(extracted_data)
            logging.info("Creating new weekly report file.")

        df = df[['Download_Date', 'PO_Number', 'Item_SKU', 'Quantity', 'Description', 'Source_File']]
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=f'Wk_{current_week}')
            
            workbook = writer.book
            worksheet = writer.sheets[f'Wk_{current_week}']
            
            # Styles
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Apply Styles to Header
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center_align, thin_border
            
            # Apply Styles to Rows
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = left_align if col in [5, 6] else center_align

            # Auto-filter and Column Widths
            worksheet.auto_filter.ref = worksheet.dimensions
            widths = {'A': 18, 'B': 15, 'C': 20, 'D': 10, 'E': 45, 'F': 25}
            for col_letter, width in widths.items():
                worksheet.column_dimensions[col_letter].width = width

        logging.info(f"Report update finalized: {output_file}")

# 5. FRIDAY ARCHIVE PROTOCOL (The Clean Desk Policy)
is_friday_late = now.weekday() == 4 and (now.hour > 16 or (now.hour == 16 and now.minute >= 45))
is_weekend = now.weekday() > 4

if is_friday_late or is_weekend:
    current_archive_path = os.path.join(archive_base, f"Week_{current_week}")
    os.makedirs(current_archive_path, exist_ok=True)
    
    pdfs_to_move = [f for f in os.listdir(base_path) if f.endswith('.pdf')]
    if pdfs_to_move:
        logging.info("Closing business week. Moving PDF files to archive...")
        for file_name in pdfs_to_move:
            try:
                shutil.move(os.path.join(base_path, file_name), os.path.join(current_archive_path, file_name))
                logging.info(f"Moved: {file_name}")
            except Exception as e:
                logging.error(f"Move error for {file_name}: {e}")
    else:
        logging.info("Archive step: Workspace already clean.")

logging.info("=== PIPELINE EXECUTION TERMINATED ===")
