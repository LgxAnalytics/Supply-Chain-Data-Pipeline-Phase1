import os
import pandas as pd
import re
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================
# 1. SETUP & PATH CONFIGURATION
# ==========================================
current_dir = os.path.dirname(os.path.abspath(__file__))
input_file_name = "Bin Contents List.xlsx" 
output_file_name = "CLEAN_MASTER_STOCK.xlsx"

input_path = os.path.join(current_dir, input_file_name)
output_path = os.path.join(current_dir, output_file_name)

# Enterprise Module: Time Machine (Archive Directory)
archive_dir = os.path.join(current_dir, "Archive")
os.makedirs(archive_dir, exist_ok=True)

def clean_stock():
    print(f"--- STARTING STOCK SANITIZATION: ENTERPRISE ENGINE ---")
    
    if not os.path.exists(input_path):
        print(f"CRITICAL ERROR: Source file not found: {input_path}")
        return

    # Load raw data
    df = pd.read_excel(input_path)
    original_df = df.copy() # Store original for the Trash Ledger
    
    sku_col = 'Item No.' 
    bin_col = 'Bin Code'
    qty_col = 'Quantity'
    loc_col = 'Location Code'

    initial_count = len(df)

    # ==========================================
    # 2. BRUTE-FORCE DATA CLEANING
    # ==========================================
    # Strip whitespaces to prevent bypass of length filters
    df[sku_col] = df[sku_col].astype(str).str.strip()
    df[loc_col] = df[loc_col].astype(str).str.strip()
    df[bin_col] = df[bin_col].astype(str).str.strip()

    # ==========================================
    # 3. FILTERING: ANOMALY DETECTION
    # ==========================================
    # Filter A: Remove short system codes (e.g., DU, CCC, DES)
    df = df[df[sku_col].str.len() >= 5]

    # Filter B: Remove non-physical assets and financial charges
    exclude_patterns = [
        'DEMO', 'HIRE', 'X-', 'WH-HIRE', 'VIRTUAL', 
        'TEST', 'SCRAP', 'DELIVERY', 'DROP SHIP', 'DEL ',
        'PACK', 'BULK'
    ]
    df = df[~df[sku_col].str.upper().str.contains('|'.join(exclude_patterns), na=False)]
    df = df[~df[loc_col].str.upper().str.contains('|'.join(exclude_patterns), na=False)]
    
    # Filter C: Keep only positive physical inventory
    df = df[df[qty_col] > 0]

    # ==========================================
    # 4. ENTERPRISE MODULE: THE TRASH LEDGER
    # ==========================================
    # Identify what was removed and save it as an audit trail (Proof of Work)
    clean_indices = df.index
    deleted_df = original_df.drop(clean_indices)
    trash_path = os.path.join(current_dir, "DELETED_ANOMALIES_LOG.xlsx")
    if not deleted_df.empty:
        deleted_df.to_excel(trash_path, index=False)
        print(f"Trash Ledger generated: {len(deleted_df)} anomalies isolated.")

    # ==========================================
    # 5. PATTERN RECOGNITION: SKU UNIFICATION (SNIPER MODE)
    # ==========================================
    def sanitize_sku(sku):
        if pd.isna(sku): return sku
        s = str(sku).upper().strip()
        # Precision cuts: Remove specific packaging flags or version modifiers
        s = re.sub(r'(-ASM|-R)$', '', s)
        s = re.sub(r'_[A-Z0-9]+$', '', s)
        return s

    df['Clean_SKU'] = df[sku_col].apply(sanitize_sku)

    # ==========================================
    # 6. HYBRID DATA AGGREGATION
    # ==========================================
    # Sum quantities while concatenating unique bin locations
    final_stock = df.groupby(['Location Code', 'Clean_SKU']).agg({
        qty_col: 'sum',
        bin_col: lambda x: ', '.join(sorted(set(x)))
    }).reset_index()
    
    # Cast quantities to absolute integers
    final_stock[qty_col] = final_stock[qty_col].astype(int)
    final_stock.rename(columns={bin_col: 'Bin_Locations'}, inplace=True)

    # ==========================================
    # 7. ENTERPRISE MODULE: DATA STAMP
    # ==========================================
    # Inject current timestamp so Power BI users know data freshness
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    final_stock['Last_Refresh'] = current_time

    # ==========================================
    # 8. EXPORT & VISUAL STYLING
    # ==========================================
    print("Exporting styled Excel outputs...")
    
    # We export the main file, and a timestamped copy for the Time Machine
    archive_path = os.path.join(archive_dir, f"CLEAN_STOCK_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    
    for out_path in [output_path, archive_path]:
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            final_stock.to_excel(writer, index=False, sheet_name='Cleaned_Stock')
            
            workbook = writer.book
            worksheet = writer.sheets['Cleaned_Stock']
            
            # Corporate UI formatting
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')

            for col in range(1, len(final_stock.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center_align, thin_border
                
            for row in range(2, len(final_stock) + 2):
                for col in range(1, len(final_stock.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    if col == 3: # Quantity
                        cell.alignment = center_align
                        cell.number_format = '#,##0'
                    elif col == 4: # Bin_Locations
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align

            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 45 # Location List
            worksheet.column_dimensions['E'].width = 20 # Timestamp

    print(f"--- SUCCESS: PIPELINE EXECUTED ---")
    print(f"Final Unique SKUs: {len(final_stock)}")
    print(f"Master Output: {output_path}")
    print(f"Archived to: {archive_path}")

if __name__ == "__main__":
    clean_stock()
